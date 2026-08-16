from __future__ import annotations

# Rich standalone HTML templates are intentionally kept inline so every report is self-contained.
# ruff: noqa: E501
import csv
import hashlib
import html
import io
import json
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from zubepredict_core.evidence import EvidenceEnvelope

REPORT_VERSION = 4

_HTML_CSP_META = (
    '<meta http-equiv="Content-Security-Policy" '
    'content="default-src \'none\'; style-src \'unsafe-inline\'; img-src data:; '
    'base-uri \'none\'; form-action \'none\'">'
)

_INK = colors.HexColor("#17233C")
_MUTED = colors.HexColor("#5F6B7A")
_NAVY = colors.HexColor("#142B4A")
_BLUE = colors.HexColor("#2474E5")
_PALE_BLUE = colors.HexColor("#EAF3FF")
_GREEN = colors.HexColor("#147D64")
_PALE_GREEN = colors.HexColor("#E8F7F1")
_AMBER = colors.HexColor("#A46100")
_PALE_AMBER = colors.HexColor("#FFF4D9")
_BORDER = colors.HexColor("#D9E1EB")
_SURFACE = colors.HexColor("#F5F8FC")

_METRIC_GUIDANCE = {
    "accuracy": "Share of predictions that were correct in validation.",
    "average_precision": "Summarises precision and recall across decision thresholds; higher is better.",
    "balanced_accuracy": "Accuracy balanced across classes, useful when classes are uneven.",
    "brier_score": "Average probability error; lower is better.",
    "f1": "Balance between precision and recall; higher is better.",
    "f1_macro": "Average F1 across classes, giving each class equal weight.",
    "log_loss": "Penalty for incorrect probabilities, especially confident mistakes; lower is better.",
    "mae": "Average absolute prediction error; lower is better.",
    "pr_auc": "Area under the precision-recall curve; higher is better.",
    "r2": "Share of outcome variation explained on validation data; higher is better.",
    "recall": "Share of actual positive cases identified; higher is better.",
    "rmse": "Typical prediction error with larger mistakes penalised more; lower is better.",
    "roc_auc": "Ability to rank positive cases above negative cases; higher is better.",
    "silhouette_score": "How well-separated the discovered clusters are; higher is better.",
}


@dataclass(frozen=True)
class GeneratedReport:
    report_type: str
    filename: str
    content_type: str
    content: bytes
    sha256: str
    report_version: int = REPORT_VERSION


def _json_bytes(value: Any) -> bytes:
    return (json.dumps(value, sort_keys=True, indent=2, default=str) + "\n").encode("utf-8")


def _report(report_type: str, filename: str, content_type: str, content: bytes) -> GeneratedReport:
    return GeneratedReport(
        report_type=report_type,
        filename=filename,
        content_type=content_type,
        content=content,
        sha256=hashlib.sha256(content).hexdigest(),
    )


def _common_payload(evidence: EvidenceEnvelope) -> dict[str, Any]:
    return {
        "title": "EyeCare Evidence Card",
        "report_version": REPORT_VERSION,
        "integrity_reference": evidence.evidence_hash,
        "experiment_id": str(evidence.experiment_id),
        "dataset_fingerprint": evidence.dataset_fingerprint,
        "constitution_version": evidence.constitution_version,
        "task": evidence.task_type,
        "target": evidence.target,
        "exclusions": evidence.exclusions,
        "validation_strategy": evidence.validation_strategy,
        "primary_metric": evidence.primary_metric,
        "secondary_metrics": evidence.secondary_metrics,
        "leaderboard": evidence.model_leaderboard,
        "selected_model": evidence.winner,
        "calibration_error_analysis": evidence.calibration_error_analysis,
        "limitations": evidence.limitations,
        "intended_use_warning": evidence.intended_use_warning,
        "reproducibility": evidence.reproducibility,
        "warnings": evidence.warnings,
        "generated_at": evidence.generated_at,
    }


def _safe(value: Any) -> str:
    return html.escape(str(value if value not in (None, "") else "Not specified"))


def _label(value: Any) -> str:
    text = str(value if value not in (None, "") else "Not specified")
    return text.replace("_", " ").replace("-", " ").strip().title()


def _number(value: Any) -> str:
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if abs(value) >= 1000:
            return f"{value:,.2f}"
        return f"{value:.3f}"
    if value in (None, ""):
        return "Not recorded"
    return str(value)


def _metric_stats(value: Any) -> dict[str, str]:
    if isinstance(value, dict):
        confidence = value.get("confidence_interval_95")
        if isinstance(confidence, (list, tuple)) and len(confidence) == 2:
            confidence_text = f"{_number(confidence[0])} to {_number(confidence[1])}"
        else:
            confidence_text = "Not recorded"
        return {
            "value": _number(value.get("mean", value.get("value"))),
            "spread": _number(value.get("standard_deviation")),
            "confidence": confidence_text,
        }
    return {"value": _number(value), "spread": "Not recorded", "confidence": "Not recorded"}


def _winner_row(payload: dict[str, Any]) -> dict[str, Any]:
    selected = payload.get("selected_model")
    for item in payload.get("leaderboard", []):
        if isinstance(item, dict) and item.get("model_name") == selected:
            return item
    for item in payload.get("leaderboard", []):
        if isinstance(item, dict):
            return item
    return {}


def _winner_metrics(payload: dict[str, Any]) -> dict[str, Any]:
    metrics = _winner_row(payload).get("metrics", {})
    return metrics if isinstance(metrics, dict) else {}


def _primary_stats(payload: dict[str, Any]) -> dict[str, str]:
    return _metric_stats(_winner_metrics(payload).get(str(payload.get("primary_metric"))))


def _generated_label(value: Any) -> str:
    if hasattr(value, "strftime"):
        return str(value.strftime("%d %B %Y at %H:%M UTC"))
    return str(value or "Not recorded")


def _list_items(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    if value in (None, "", {}):
        return []
    return [str(value)]


def _executive_summary(payload: dict[str, Any]) -> str:
    stats = _primary_stats(payload)
    return (
        f"The selected model was {_label(payload.get('selected_model'))}. "
        f"Its recorded {_label(payload.get('primary_metric')).lower()} was {stats['value']} "
        f"under {_label(payload.get('validation_strategy')).lower()}. "
        "This result describes the recorded validation exercise and does not guarantee future performance."
    )


def _plain_validation(value: Any) -> str:
    text = str(value or "Not recorded")
    replacements = {
        "StratifiedKFold": "stratified cross-validation",
        "KFold": "cross-validation",
        "fold-fitted": "separately fitted within each test fold",
        "fold fitted": "separately fitted within each test fold",
    }
    for original, replacement in replacements.items():
        text = text.replace(original, replacement)
    return text


def _metric_direction(name: str) -> str:
    return "Lower values are better." if name in {"brier_score", "log_loss", "mae", "rmse"} else "Higher values are better."


def _html_report(payload: dict[str, Any]) -> bytes:
    primary = _primary_stats(payload)
    winner_metrics = _winner_metrics(payload)
    primary_name = str(payload.get("primary_metric"))
    secondary_rows = "".join(
        f"<tr><th scope=\"row\">{_safe(_label(name))}</th>"
        f"<td>{_safe(stats['value'])}</td><td>{_safe(stats['spread'])}</td>"
        f"<td>{_safe(stats['confidence'])}</td>"
        f"<td>{_safe(_METRIC_GUIDANCE.get(str(name), 'Recorded validation measure.'))}</td></tr>"
        for name, value in winner_metrics.items()
        if name != primary_name
        for stats in [_metric_stats(value)]
    )
    leaderboard_parts: list[str] = []
    for rank, item in enumerate(payload.get("leaderboard", []), start=1):
        if not isinstance(item, dict):
            continue
        selected = item.get("model_name") == payload.get("selected_model")
        marker = '<span class="winner-tag">Selected</span>' if selected else ""
        row_class = "selected-row" if selected else ""
        leaderboard_parts.append(
            f'<tr class="{row_class}"><td>{rank}</td>'
            f'<th scope="row">{_safe(_label(item.get("model_name")))}{marker}</th>'
            f'<td>{_safe(_metric_stats((item.get("metrics") or {}).get(primary_name))["value"])}</td>'
            f'<td>{_safe(_label(item.get("status", "completed")))}</td></tr>'
        )
    leaderboard_rows = "".join(leaderboard_parts)
    exclusions = "".join(f"<span class=\"chip\">{_safe(_label(item))}</span>" for item in payload["exclusions"])
    limitations = "".join(f"<li>{_safe(item)}</li>" for item in _list_items(payload["limitations"]))
    warnings = "".join(f"<li>{_safe(item)}</li>" for item in _list_items(payload["warnings"]))
    analysis = payload.get("calibration_error_analysis", {})
    calibration = analysis.get("calibration", {}) if isinstance(analysis, dict) else {}
    threshold = analysis.get("threshold_analysis", {}) if isinstance(analysis, dict) else {}
    error_summary = analysis.get("error_analysis_summary", {}) if isinstance(analysis, dict) else {}
    analysis_cards = "".join(
        f"<div class=\"mini-card\"><span>{_safe(title)}</span><strong>{_safe(value)}</strong></div>"
        for title, value in (
            ("Brier score", _number(calibration.get("brier_score"))),
            ("Calibration error", _number(calibration.get("expected_calibration_error"))),
            ("Recommended threshold", _number(threshold.get("recommended_threshold"))),
            ("Segments reviewed", _number(error_summary.get("segment_count"))),
        )
    )
    software = payload.get("reproducibility", {}).get("software_versions", {})
    software_rows = "".join(
        f"<tr><th scope=\"row\">{_safe(_label(name))}</th><td>{_safe(version)}</td></tr>"
        for name, version in software.items()
    )
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">{_HTML_CSP_META}
<title>ZubePredict AI Evidence Report</title>
<style>
:root{{--ink:#17233c;--muted:#667085;--navy:#142b4a;--blue:#2474e5;--line:#d9e1eb;--surface:#f5f8fc;--green:#147d64;--green-bg:#e8f7f1;--amber:#8a5200;--amber-bg:#fff4d9}}
*{{box-sizing:border-box}} html{{scroll-behavior:smooth}} body{{margin:0;overflow-x:hidden;background:#e9f0f7;color:var(--ink);font:15px/1.65 Inter,Segoe UI,Arial,sans-serif}}p,h1,h2,h3,th,td,dd,strong{{overflow-wrap:anywhere}}
.report{{width:min(1080px,calc(100% - 32px));margin:32px auto;background:#fff;border-radius:22px;overflow:hidden;box-shadow:0 22px 70px #17304f24}}
.hero{{padding:52px 58px 44px;background:linear-gradient(135deg,#102743,#173c67 70%,#2474e5);color:#fff;position:relative;overflow:hidden}}
.hero:after{{content:"";position:absolute;width:280px;height:280px;border:50px solid #ffffff12;border-radius:50%;right:-90px;top:-110px}}
.brand{{display:flex;align-items:center;gap:10px;font-size:14px;font-weight:800;letter-spacing:.09em;text-transform:uppercase}} .brand-mark{{display:grid;place-items:center;width:32px;height:32px;border-radius:10px;background:#fff;color:#173c67;font-size:17px;letter-spacing:0}} .eyebrow{{margin:35px 0 6px;color:#bcd8ff;font-size:12px;font-weight:800;letter-spacing:.12em;text-transform:uppercase}}
h1{{font-size:42px;line-height:1.08;margin:0 0 14px}} .hero p{{max-width:720px;margin:0;color:#e8f1fc;font-size:17px}}
.verified{{display:inline-flex;margin-top:24px;padding:7px 12px;border:1px solid #8be1c5;border-radius:99px;background:#0c765c;color:#fff;font-weight:700;font-size:12px;letter-spacing:.05em}}
main{{padding:42px 58px 54px}} h2{{margin:0 0 8px;font-size:25px;color:var(--navy)}} h3{{margin:0 0 6px;font-size:17px}} .lead{{font-size:17px;color:#344054;margin:0}}
.summary{{padding:24px;border-left:5px solid var(--blue);background:#f0f6ff;border-radius:0 12px 12px 0;margin-bottom:28px}}
.cards{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:24px 0 34px}} .card,.mini-card{{border:1px solid var(--line);border-radius:12px;padding:17px;background:#fff}}
.card span,.mini-card span{{display:block;color:var(--muted);font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.04em}} .card strong{{display:block;margin-top:6px;font-size:19px;color:var(--navy)}}
.reading-guide{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:0 0 34px}}.guide{{position:relative;padding:20px 18px 18px 54px;border:1px solid var(--line);border-radius:14px;background:var(--surface)}}.guide>b{{position:absolute;left:17px;top:18px;display:grid;place-items:center;width:26px;height:26px;border-radius:50%;background:var(--blue);color:#fff}}.guide strong{{display:block;color:var(--navy)}}.guide p{{margin:4px 0 0;color:var(--muted);font-size:13px;line-height:1.5}}
section{{padding:28px 0;border-top:1px solid var(--line)}} table{{border-collapse:separate;border-spacing:0;width:100%;margin-top:16px;border:1px solid var(--line);border-radius:10px;overflow:hidden}}
th,td{{padding:12px 14px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}} thead th{{background:var(--navy);color:#fff;font-size:12px;text-transform:uppercase;letter-spacing:.04em}} tbody th{{color:var(--navy)}} tr:last-child td,tr:last-child th{{border-bottom:0}}
.selected-row td,.selected-row th{{background:var(--green-bg)}}.winner-tag{{display:inline-block;margin-left:9px;padding:2px 7px;border-radius:99px;background:var(--green);color:#fff;font-size:10px;letter-spacing:.04em;text-transform:uppercase;vertical-align:middle}}
.warning{{padding:20px 22px;border:1px solid #f0d08a;background:var(--amber-bg);color:#603c00;border-radius:12px;margin:26px 0}} .warning strong{{display:block;color:var(--amber);margin-bottom:4px}}
.explanation{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:18px}}.explain-box{{padding:18px;border-radius:12px;background:var(--surface);border:1px solid var(--line)}}.explain-box.good{{background:var(--green-bg);border-color:#a9ddcc}}.explain-box strong{{display:block;color:var(--navy);margin-bottom:5px}}.explain-box p{{margin:0;color:#475467}}
.chips{{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px}} .chip{{padding:6px 10px;border-radius:99px;background:#eaf3ff;color:#185cba;font-weight:700;font-size:13px}}
.analysis{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:18px}} .mini-card{{background:var(--surface)}} .mini-card strong{{display:block;margin-top:7px;font-size:17px}}
.clean-list{{margin:14px 0 0;padding-left:22px}} .clean-list li{{margin:7px 0}} code{{font:12px/1.5 Consolas,monospace;word-break:break-all;color:#344054}}
.meta{{display:grid;grid-template-columns:180px 1fr;gap:10px 18px;margin-top:16px}} .meta dt{{font-weight:700;color:var(--muted)}} .meta dd{{margin:0}}
details{{margin-top:18px;padding:15px 17px;border:1px solid var(--line);border-radius:11px;background:#fbfcfe}} summary{{cursor:pointer;font-weight:700;color:var(--navy)}}details[open] summary{{margin-bottom:14px}}
.footer{{padding:20px 58px;background:var(--surface);border-top:1px solid var(--line);color:var(--muted);font-size:12px;display:flex;justify-content:space-between;gap:20px}}
@media(max-width:760px){{.hero,main{{padding:30px 24px}}h1{{font-size:32px}}.cards,.analysis,.reading-guide,.explanation{{grid-template-columns:minmax(0,1fr) minmax(0,1fr)}}.meta{{grid-template-columns:minmax(0,1fr)}}.footer{{padding:18px 24px;display:block}}table{{display:block;max-width:100%;overflow-x:auto}}}}
@media(max-width:500px){{.cards,.analysis,.reading-guide,.explanation{{grid-template-columns:1fr}}.report{{width:100%;margin:0;border-radius:0}}}}
@media print{{body{{background:#fff}}.report{{width:100%;margin:0;box-shadow:none}}.hero{{print-color-adjust:exact;-webkit-print-color-adjust:exact}}section,table,.warning{{break-inside:avoid}}details{{display:block}}}}
</style></head><body><article class="report">
<header class="hero"><div class="brand"><span class="brand-mark">Z</span>ZubePredict AI</div><div class="eyebrow">Verified model evaluation</div><h1>Evidence Report</h1>
<p>A guided explanation of what was tested, what performed best, and what you should understand before using the result.</p><span class="verified">VERIFIED EVIDENCE &middot; VERSION {payload['report_version']}</span></header>
<main><div class="summary"><h2>What this result says</h2><p class="lead">{_safe(_executive_summary(payload))}</p></div>
<div class="cards"><div class="card"><span>Selected model</span><strong>{_safe(_label(payload['selected_model']))}</strong></div>
<div class="card"><span>{_safe(_label(primary_name))}</span><strong>{_safe(primary['value'])}</strong></div>
<div class="card"><span>Task</span><strong>{_safe(_label(payload['task']))}</strong></div>
<div class="card"><span>Target</span><strong>{_safe(_label(payload['target']))}</strong></div></div>
<div class="reading-guide"><div class="guide"><b>1</b><strong>What was predicted?</strong><p>The model was trained to predict <b>{_safe(_label(payload['target']))}</b>.</p></div><div class="guide"><b>2</b><strong>Which model was chosen?</strong><p><b>{_safe(_label(payload['selected_model']))}</b> ranked first under the approved test.</p></div><div class="guide"><b>3</b><strong>What should I do next?</strong><p>Read the limitations and validate independently before any real-world use.</p></div></div>
<div class="warning"><strong>Important use limitation</strong>{_safe(payload['intended_use_warning'])}</div>
<section><h2>Study design</h2><p>The model was evaluated using the approved Experiment Constitution and the validation plan below.</p>
<dl class="meta"><dt>How it was tested</dt><dd>{_safe(_plain_validation(payload['validation_strategy']))}</dd><dt>Main measure</dt><dd>{_safe(_label(primary_name))}: {_safe(_METRIC_GUIDANCE.get(primary_name, 'The main recorded validation measure.'))} {_safe(_metric_direction(primary_name))}</dd>
<dt>95% confidence interval</dt><dd>{_safe(primary['confidence'])}</dd><dt>Standard deviation</dt><dd>{_safe(primary['spread'])}</dd></dl>
<div class="explanation"><div class="explain-box good"><strong>What the score means</strong><p>This is the score recorded during the approved validation exercise. It compares the candidate models under the same rules.</p></div><div class="explain-box"><strong>What the score does not mean</strong><p>It is not a promise of future performance and it does not prove that the model is safe for clinical or operational decisions.</p></div></div>
<h3 style="margin-top:20px">Excluded fields</h3><div class="chips">{exclusions or '<span class="chip">None recorded</span>'}</div></section>
<section><h2>Model leaderboard</h2><p>Models are shown in the recorded order. The selected model is the authoritative winner in the Evidence Envelope.</p>
<table><thead><tr><th>Rank</th><th>Model</th><th>{_safe(_label(primary_name))}</th><th>Run status</th></tr></thead><tbody>{leaderboard_rows or '<tr><td colspan="4">No leaderboard was recorded.</td></tr>'}</tbody></table></section>
<section><h2>Supporting measures</h2><p>These measures describe different aspects of the selected model. They should be read together, not as isolated scores.</p>
<table><thead><tr><th>Measure</th><th>Mean</th><th>Std. dev.</th><th>95% interval</th><th>Plain-language meaning</th></tr></thead><tbody>{secondary_rows or '<tr><td colspan="5">No secondary measures were recorded.</td></tr>'}</tbody></table></section>
<section><h2>Calibration and error review</h2><p>Calibration asks whether predicted probabilities behave like real-world frequencies. Error review records the checks performed around mistakes and decision thresholds.</p>
<div class="analysis">{analysis_cards}</div>
<details><summary>Technical notes</summary><dl class="meta"><dt>Threshold basis</dt><dd>{_safe(threshold.get('recommendation_basis'))}</dd><dt>Calibration bins</dt><dd>{_safe(len(calibration.get('bins', [])))}</dd><dt>Plots recorded</dt><dd>{_safe(', '.join(_label(item) for item in error_summary.get('plot_ids', [])) or 'None recorded')}</dd><dt>Protected columns skipped</dt><dd>{_safe(', '.join(str(item) for item in error_summary.get('protected_columns_skipped', [])) or 'None recorded')}</dd></dl></details></section>
<section><h2>Limitations and warnings</h2><ul class="clean-list">{limitations or '<li>No additional limitations were recorded.</li>'}</ul>{f'<h3 style="margin-top:18px">Run warnings</h3><ul class="clean-list">{warnings}</ul>' if warnings else ''}</section>
<section><h2>Technical and audit information</h2><p>You do not need this section to understand the result. It is preserved for reviewers who need to reproduce or verify the experiment.</p><details><summary>Show reproducibility details</summary><dl class="meta"><dt>Random seed</dt><dd>{_safe(payload.get('reproducibility', {}).get('random_seed'))}</dd><dt>Generated</dt><dd>{_safe(_generated_label(payload['generated_at']))}</dd><dt>Report version</dt><dd>{payload['report_version']}</dd></dl>
<table><thead><tr><th>Software</th><th>Version</th></tr></thead><tbody>{software_rows or '<tr><td colspan="2">No software versions were recorded.</td></tr>'}</tbody></table></details>
<details><summary>Show integrity and traceability details</summary><dl class="meta"><dt>Experiment ID</dt><dd><code>{_safe(payload['experiment_id'])}</code></dd><dt>Dataset fingerprint</dt><dd><code>{_safe(payload['dataset_fingerprint'])}</code></dd><dt>Constitution version</dt><dd>{_safe(payload['constitution_version'])}</dd><dt>Evidence hash</dt><dd><code>{_safe(payload['integrity_reference'])}</code></dd></dl></details></section></main>
<footer class="footer"><span>ZubePredict AI &middot; Evidence, not assertion.</span><span>Decision support and research unless independently validated.</span></footer></article></body></html>"""
    return document.encode("utf-8")


def _html_evidence_card(payload: dict[str, Any]) -> bytes:
    primary = _primary_stats(payload)
    limitations = "".join(f"<li>{_safe(item)}</li>" for item in _list_items(payload["limitations"]))
    metric_name = str(payload["primary_metric"])
    document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">{_HTML_CSP_META}
<title>EyeCare Evidence Card</title><style>
:root{{--navy:#122a49;--blue:#2474e5;--ink:#17233c;--muted:#667085;--line:#d8e2ed;--surface:#f5f8fc;--green:#147d64;--green-bg:#e8f7f1;--amber:#805000;--amber-bg:#fff4d9}}*{{box-sizing:border-box}}body{{margin:0;background:#e9f0f7;color:var(--ink);font:15px/1.65 Inter,Segoe UI,Arial,sans-serif}}main{{width:min(900px,calc(100% - 32px));margin:32px auto;background:#fff;border-radius:22px;overflow:hidden;box-shadow:0 22px 70px #17304f24}}header{{padding:42px 48px;background:linear-gradient(135deg,#102743,#174675 72%,#2474e5);color:#fff;position:relative;overflow:hidden}}header:after{{content:"";position:absolute;width:220px;height:220px;border:42px solid #ffffff12;border-radius:50%;right:-80px;top:-100px}}.brand{{display:flex;align-items:center;gap:10px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}}.mark{{display:grid;place-items:center;width:32px;height:32px;border-radius:10px;background:#fff;color:var(--navy)}}.eyebrow{{margin-top:30px;color:#bcd8ff;font-size:12px;font-weight:800;letter-spacing:.11em;text-transform:uppercase}}h1{{margin:5px 0 8px;font-size:38px;line-height:1.15}}header p{{margin:0;max-width:650px;color:#e7f1ff;font-size:16px}}.stamp{{display:inline-block;margin-top:20px;padding:6px 11px;border:1px solid #8be1c5;border-radius:99px;background:#0c765c;font-size:11px;font-weight:800;letter-spacing:.05em}}.body{{padding:38px 48px 44px}}.intro{{margin:0 0 24px;color:#475467;font-size:16px}}.result{{display:grid;grid-template-columns:1fr auto;gap:24px;align-items:center;padding:24px 26px;border:1px solid #afd5ff;background:#edf6ff;border-radius:16px}}.result h2{{margin:0 0 5px;font-size:21px;color:var(--navy)}}.result p{{margin:0;color:#475467}}.score{{min-width:155px;padding:16px;text-align:center;border-radius:13px;background:#fff;box-shadow:0 6px 20px #17304f12}}.score span{{display:block;color:var(--muted);font-size:11px;font-weight:800;text-transform:uppercase}}.score strong{{display:block;color:var(--blue);font-size:32px;line-height:1.2}}.grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:24px 0}}.item{{border:1px solid var(--line);border-radius:13px;padding:17px;background:#fff}}.item span{{display:block;color:var(--muted);font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.04em}}.item strong{{display:block;margin-top:6px;font-size:17px;color:var(--navy)}}h2.section{{font-size:21px;margin:30px 0 10px;color:var(--navy)}}.steps{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}}.step{{padding:17px;border-radius:13px;background:var(--surface);border:1px solid var(--line)}}.step b{{display:grid;place-items:center;width:25px;height:25px;margin-bottom:10px;border-radius:50%;background:var(--blue);color:#fff}}.step strong{{display:block;color:var(--navy)}}.step p{{margin:4px 0 0;color:var(--muted);font-size:13px;line-height:1.5}}.warning{{margin-top:26px;padding:20px 22px;background:var(--amber-bg);border:1px solid #efd087;border-radius:13px;color:#603c00}}.warning strong{{display:block;color:var(--amber);margin-bottom:4px}}li{{margin:7px 0}}details{{margin-top:25px;padding:15px 17px;border:1px solid var(--line);border-radius:11px;background:#fbfcfe}}summary{{cursor:pointer;color:var(--navy);font-weight:800}}.trace{{display:grid;grid-template-columns:160px 1fr;gap:10px 16px;margin-top:15px}}.trace dt{{color:var(--muted);font-weight:700}}.trace dd{{margin:0}}code{{font:12px/1.5 Consolas,monospace;word-break:break-all}}footer{{padding:19px 48px;background:var(--surface);border-top:1px solid var(--line);color:var(--muted);font-size:12px}}@media(max-width:680px){{header,.body{{padding:30px 23px}}.result{{grid-template-columns:1fr}}.score{{text-align:left}}.grid,.steps{{grid-template-columns:1fr}}.trace{{grid-template-columns:1fr}}}}@media(max-width:480px){{main{{width:100%;margin:0;border-radius:0}}h1{{font-size:31px}}}}
body{{overflow-x:hidden}}p,h1,h2,dt,dd,strong{{overflow-wrap:anywhere}}.result>*,.grid>*,.steps>*{{min-width:0}}
</style></head><body><main><header><div class="brand"><span class="mark">Z</span>ZubePredict AI</div><div class="eyebrow">A short, verified summary</div><h1>EyeCare Evidence Card</h1><p>The most important findings and cautions from one completed model experiment.</p><span class="stamp">VERIFIED &middot; VERSION {payload['report_version']}</span></header><div class="body">
<p class="intro"><strong>Start here.</strong> This card gives you the short answer. Use the full Evidence Report when you need the complete model comparison and testing details.</p>
<div class="result"><div><h2>What the experiment selected</h2><p><strong>{_safe(_label(payload['selected_model']))}</strong> was selected to predict <strong>{_safe(_label(payload['target']))}</strong>. This reflects the recorded validation exercise, not a guarantee about future cases.</p></div><div class="score"><span>{_safe(_label(metric_name))}</span><strong>{_safe(primary['value'])}</strong><small>{_safe(_metric_direction(metric_name))}</small></div></div>
<div class="grid"><div class="item"><span>Prediction type</span><strong>{_safe(_label(payload['task']))}</strong></div><div class="item"><span>What is predicted</span><strong>{_safe(_label(payload['target']))}</strong></div><div class="item"><span>How it was tested</span><strong>{_safe(_plain_validation(payload['validation_strategy']))}</strong></div></div>
<h2 class="section">How to read this card</h2><div class="steps"><div class="step"><b>1</b><strong>Result</strong><p>The selected model performed best under the approved comparison rules.</p></div><div class="step"><b>2</b><strong>Caution</strong><p>The recorded score may not repeat on new people, places, or time periods.</p></div><div class="step"><b>3</b><strong>Next step</strong><p>Validate independently before using the model for real decisions.</p></div></div>
<div class="warning"><strong>Intended use</strong>{_safe(payload['intended_use_warning'])}</div><h2 class="section">Important limitations</h2><ul>{limitations or '<li>No additional limitations were recorded.</li>'}</ul>
<details><summary>Show technical traceability details</summary><dl class="trace"><dt>Experiment</dt><dd><code>{_safe(payload['experiment_id'])}</code></dd><dt>Dataset fingerprint</dt><dd><code>{_safe(payload['dataset_fingerprint'])}</code></dd><dt>Evidence hash</dt><dd><code>{_safe(payload['integrity_reference'])}</code></dd><dt>Generated</dt><dd>{_safe(_generated_label(payload['generated_at']))}</dd></dl></details></div><footer>This card summarises the immutable Evidence Envelope. Values are generated by the backend, not rewritten by an AI assistant.</footer></main></body></html>"""
    return document.encode("utf-8")


def _model_card(payload: dict[str, Any]) -> bytes:
    primary_name = str(payload["primary_metric"])
    primary = _primary_stats(payload)
    reproducibility = payload.get("reproducibility", {})
    software = reproducibility.get("software_versions", {})
    hyperparameters = reproducibility.get("winner_hyperparameters", {})
    limitations = "".join(f"<li>{_safe(item)}</li>" for item in _list_items(payload["limitations"]))
    software_rows = "".join(
        f"<tr><th>{_safe(_label(name))}</th><td>{_safe(value)}</td></tr>"
        for name, value in software.items()
    )
    parameter_rows = "".join(
        f"<tr><th>{_safe(_label(name))}</th><td>{_safe(_number(value))}</td></tr>"
        for name, value in hyperparameters.items()
    )
    document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">{_HTML_CSP_META}<title>ZubePredict AI Model Card</title>
<style>:root{{--navy:#122a49;--blue:#2474e5;--ink:#17233c;--muted:#667085;--line:#d8e2ed;--surface:#f5f8fc;--green:#147d64;--green-bg:#e8f7f1;--amber:#805000;--amber-bg:#fff4d9}}*{{box-sizing:border-box}}body{{margin:0;background:#e9f0f7;color:var(--ink);font:15px/1.65 Inter,Segoe UI,Arial,sans-serif}}article{{width:min(980px,calc(100% - 32px));margin:32px auto;background:#fff;border-radius:22px;overflow:hidden;box-shadow:0 22px 70px #17304f24}}header{{padding:46px 52px;background:linear-gradient(135deg,#102743,#173c67 70%,#2474e5);color:#fff;position:relative;overflow:hidden}}header:after{{content:"";position:absolute;width:240px;height:240px;border:45px solid #ffffff12;border-radius:50%;right:-90px;top:-110px}}.brand{{display:flex;align-items:center;gap:10px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}}.mark{{display:grid;place-items:center;width:32px;height:32px;border-radius:10px;background:#fff;color:var(--navy)}}.eyebrow{{margin-top:31px;color:#bcd8ff;font-size:12px;font-weight:800;letter-spacing:.11em;text-transform:uppercase}}h1{{font-size:40px;line-height:1.12;margin:5px 0 9px}}header p{{max-width:680px;margin:0;color:#e7f1ff;font-size:16px}}.stamp{{display:inline-block;margin-top:21px;padding:6px 11px;border:1px solid #8be1c5;border-radius:99px;background:#0c765c;font-size:11px;font-weight:800;letter-spacing:.05em}}main{{padding:40px 52px 50px}}.what-is{{padding:18px 20px;border-radius:13px;background:var(--surface);border:1px solid var(--line);color:#475467}}.what-is strong{{color:var(--navy)}}.result{{display:grid;grid-template-columns:1fr auto;align-items:center;gap:24px;margin-top:24px;padding:25px 27px;border-radius:16px;background:#edf6ff;border:1px solid #afd5ff}}.result h2{{margin:0 0 5px;font-size:22px}}.result p{{margin:0;color:#475467}}.score{{min-width:155px;padding:16px;text-align:center;border-radius:13px;background:#fff;box-shadow:0 6px 20px #17304f12}}.score span{{display:block;color:var(--muted);font-size:11px;font-weight:800;text-transform:uppercase}}.score strong{{display:block;color:var(--blue);font-size:32px;line-height:1.2}}.facts{{display:grid;grid-template-columns:repeat(3,1fr);gap:13px;margin:22px 0 32px}}.fact{{padding:17px;border:1px solid var(--line);border-radius:12px}}.fact span{{display:block;color:var(--muted);font-size:11px;font-weight:800;text-transform:uppercase}}.fact strong{{display:block;margin-top:6px;color:var(--navy);font-size:17px}}section{{padding:28px 0;border-top:1px solid var(--line)}}section h2{{margin:0 0 8px;font-size:23px;color:var(--navy)}}section p{{margin:0;color:#475467}}.plain-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:17px}}.plain{{padding:18px;border-radius:13px;background:var(--surface);border:1px solid var(--line)}}.plain strong{{display:block;margin-bottom:5px;color:var(--navy)}}.plain.good{{background:var(--green-bg);border-color:#a9ddcc}}.warning{{margin-top:16px;padding:20px 22px;background:var(--amber-bg);border:1px solid #efd087;border-radius:13px;color:#603c00}}.warning strong{{display:block;color:var(--amber);margin-bottom:4px}}li{{margin:7px 0}}details{{margin-top:14px;padding:15px 17px;border:1px solid var(--line);border-radius:11px;background:#fbfcfe}}summary{{cursor:pointer;color:var(--navy);font-weight:800}}details[open] summary{{margin-bottom:14px}}table{{border-collapse:separate;border-spacing:0;width:100%;border:1px solid var(--line);border-radius:10px;overflow:hidden}}th,td{{padding:11px 13px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}}tr:last-child th,tr:last-child td{{border-bottom:0}}th{{width:34%;background:var(--surface);color:var(--navy)}}code{{font:12px/1.5 Consolas,monospace;word-break:break-all}}footer{{padding:19px 52px;background:var(--surface);border-top:1px solid var(--line);color:var(--muted);font-size:12px}}@media(max-width:700px){{header,main{{padding:30px 23px}}.result{{grid-template-columns:1fr}}.score{{text-align:left}}.facts,.plain-grid{{grid-template-columns:1fr}}table{{display:block;overflow-x:auto}}}}@media(max-width:480px){{article{{width:100%;margin:0;border-radius:0}}h1{{font-size:32px}}}}</style></head>
<style>body{{overflow-x:hidden}}p,h1,h2,th,td,strong{{overflow-wrap:anywhere}}.result>*,.facts>*,.plain-grid>*{{min-width:0}}</style>
<body><article><header><div class="brand"><span class="mark">Z</span>ZubePredict AI</div><div class="eyebrow">Plain-language model documentation</div><h1>Model Card</h1><p>What the selected model does, how it was tested, when it may help, and where caution is required.</p><span class="stamp">VERIFIED &middot; VERSION {payload['report_version']}</span></header><main>
<div class="what-is"><strong>What is a Model Card?</strong> It is a factsheet for a trained model. It explains the model's purpose and boundaries so that someone can review it before deciding whether it deserves further testing.</div>
<div class="result"><div><h2>Selected model: {_safe(_label(payload['selected_model']))}</h2><p>This model was selected from the recorded candidates to predict <strong>{_safe(_label(payload['target']))}</strong>. Selection was based on the approved primary measure.</p></div><div class="score"><span>{_safe(_label(primary_name))}</span><strong>{_safe(primary['value'])}</strong><small>{_safe(_metric_direction(primary_name))}</small></div></div>
<div class="facts"><div class="fact"><span>Prediction type</span><strong>{_safe(_label(payload['task']))}</strong></div><div class="fact"><span>What it predicts</span><strong>{_safe(_label(payload['target']))}</strong></div><div class="fact"><span>Confidence interval</span><strong>{_safe(primary['confidence'])}</strong></div></div>
<section><h2>What this model does</h2><p>It looks for patterns in the approved dataset and produces a prediction for <strong>{_safe(_label(payload['target']))}</strong>. A prediction is decision support; it is not proof, diagnosis, or causation.</p><div class="plain-grid"><div class="plain good"><strong>How it was evaluated</strong>{_safe(_plain_validation(payload['validation_strategy']))}</div><div class="plain"><strong>How to read the main score</strong>{_safe(_METRIC_GUIDANCE.get(primary_name, 'This is the main recorded validation measure.'))} {_safe(_metric_direction(primary_name))}</div></div></section>
<section><h2>Responsible use</h2><p>Review these boundaries before the model is used outside this experiment.</p><div class="warning"><strong>Intended-use warning</strong>{_safe(payload['intended_use_warning'])}</div></section>
<section><h2>Known limitations</h2><ul>{limitations or '<li>No additional limitations were recorded.</li>'}</ul><p>The recorded results describe this dataset and this validation design. Performance can change when the population, data collection process, or time period changes.</p></section>
<section><h2>Technical appendix</h2><p>These details are primarily for technical reviewers and auditors.</p><details><summary>Show validation details</summary><table><tr><th>Validation strategy</th><td>{_safe(payload['validation_strategy'])}</td></tr><tr><th>Primary measure</th><td>{_safe(_label(primary_name))}: {_safe(primary['value'])}</td></tr><tr><th>95% interval</th><td>{_safe(primary['confidence'])}</td></tr><tr><th>Excluded fields</th><td>{_safe(', '.join(payload['exclusions']) or 'None recorded')}</td></tr></table></details>
<details><summary>Show software versions</summary><table>{software_rows or '<tr><td>Not recorded</td></tr>'}</table></details><details><summary>Show selected model settings</summary><table>{parameter_rows or '<tr><td>Not recorded</td></tr>'}</table></details>
<details><summary>Show integrity and traceability</summary><table><tr><th>Experiment ID</th><td><code>{_safe(payload['experiment_id'])}</code></td></tr><tr><th>Dataset fingerprint</th><td><code>{_safe(payload['dataset_fingerprint'])}</code></td></tr><tr><th>Evidence hash</th><td><code>{_safe(payload['integrity_reference'])}</code></td></tr><tr><th>Report version</th><td>{payload['report_version']}</td></tr></table></details></section></main><footer>Decision support and research unless independently validated. The Evidence Envelope remains the authoritative source.</footer></article></body></html>"""
    document = document.replace("</style></head>\n<style>body", "</style>\n<style>body", 1)
    document = document.replace("</style>\n<body>", "</style></head>\n<body>", 1)
    return document.encode("utf-8")


def _html_reproducibility_manifest(payload: dict[str, Any], artifact_policy: str) -> bytes:
    reproducibility = payload.get("reproducibility", {})
    software = reproducibility.get("software_versions", {})
    hyperparameters = reproducibility.get("winner_hyperparameters", {})
    software_rows = "".join(
        f"<tr><th scope=\"row\">{_safe(_label(name))}</th><td>{_safe(version)}</td></tr>"
        for name, version in software.items()
    )
    parameter_rows = "".join(
        f"<tr><th scope=\"row\">{_safe(_label(name))}</th><td>{_safe(_number(value))}</td></tr>"
        for name, value in hyperparameters.items()
    )
    exclusions = "".join(
        f'<span class="chip">{_safe(_label(item))}</span>' for item in payload["exclusions"]
    )
    document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">{_HTML_CSP_META}<title>ZubePredict AI Reproducibility Manifest</title>
<style>:root{{--navy:#122a49;--blue:#2474e5;--ink:#17233c;--muted:#667085;--line:#d8e2ed;--surface:#f5f8fc;--green:#147d64;--green-bg:#e8f7f1;--amber:#805000;--amber-bg:#fff4d9}}*{{box-sizing:border-box}}body{{margin:0;overflow-x:hidden;background:#e9f0f7;color:var(--ink);font:15px/1.65 Inter,Segoe UI,Arial,sans-serif}}p,h1,h2,th,td,strong,dd{{overflow-wrap:anywhere}}article{{width:min(960px,calc(100% - 32px));margin:32px auto;background:#fff;border-radius:22px;overflow:hidden;box-shadow:0 22px 70px #17304f24}}header{{padding:46px 52px;background:linear-gradient(135deg,#102743,#173c67 70%,#2474e5);color:#fff;position:relative;overflow:hidden}}header:after{{content:"";position:absolute;width:240px;height:240px;border:45px solid #ffffff12;border-radius:50%;right:-90px;top:-110px}}.brand{{display:flex;align-items:center;gap:10px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}}.mark{{display:grid;place-items:center;width:32px;height:32px;border-radius:10px;background:#fff;color:var(--navy)}}.eyebrow{{margin-top:31px;color:#bcd8ff;font-size:12px;font-weight:800;letter-spacing:.11em;text-transform:uppercase}}h1{{font-size:38px;line-height:1.12;margin:5px 0 9px}}header p{{max-width:680px;margin:0;color:#e7f1ff;font-size:16px}}.stamp{{display:inline-block;margin-top:21px;padding:6px 11px;border:1px solid #8be1c5;border-radius:99px;background:#0c765c;font-size:11px;font-weight:800;letter-spacing:.05em}}main{{padding:40px 52px 50px}}.intro{{padding:20px 22px;border:1px solid #afd5ff;border-radius:14px;background:#edf6ff;color:#344054}}.intro strong{{display:block;margin-bottom:4px;color:var(--navy);font-size:18px}}.steps{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:13px;margin:24px 0 32px}}.step{{padding:18px;border:1px solid var(--line);border-radius:13px;background:var(--surface)}}.step b{{display:grid;place-items:center;width:26px;height:26px;margin-bottom:10px;border-radius:50%;background:var(--blue);color:#fff}}.step strong{{display:block;color:var(--navy)}}.step p{{margin:4px 0 0;color:var(--muted);font-size:13px}}section{{padding:28px 0;border-top:1px solid var(--line)}}section h2{{margin:0 0 8px;font-size:23px;color:var(--navy)}}section>p{{margin:0;color:#475467}}.facts{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:13px;margin-top:17px}}.fact{{padding:17px;border:1px solid var(--line);border-radius:12px}}.fact span{{display:block;color:var(--muted);font-size:11px;font-weight:800;text-transform:uppercase}}.fact strong{{display:block;margin-top:5px;color:var(--navy);font-size:17px}}.policy{{margin-top:18px;padding:19px 21px;border:1px solid #a9ddcc;border-radius:13px;background:var(--green-bg);color:#285f52}}.policy strong{{display:block;color:var(--green);margin-bottom:4px}}.warning{{margin-top:18px;padding:19px 21px;border:1px solid #efd087;border-radius:13px;background:var(--amber-bg);color:#603c00}}.warning strong{{display:block;color:var(--amber);margin-bottom:4px}}.chips{{display:flex;gap:8px;flex-wrap:wrap;margin-top:13px}}.chip{{padding:6px 10px;border-radius:99px;background:#eaf3ff;color:#185cba;font-size:13px;font-weight:700}}details{{margin-top:14px;padding:15px 17px;border:1px solid var(--line);border-radius:11px;background:#fbfcfe}}summary{{cursor:pointer;color:var(--navy);font-weight:800}}details[open] summary{{margin-bottom:14px}}table{{border-collapse:separate;border-spacing:0;width:100%;border:1px solid var(--line);border-radius:10px;overflow:hidden}}th,td{{padding:11px 13px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}}tr:last-child th,tr:last-child td{{border-bottom:0}}th{{width:35%;background:var(--surface);color:var(--navy)}}code{{font:12px/1.5 Consolas,monospace;word-break:break-all}}footer{{padding:19px 52px;background:var(--surface);border-top:1px solid var(--line);color:var(--muted);font-size:12px}}@media(max-width:700px){{header,main{{padding:30px 23px}}.steps,.facts{{grid-template-columns:minmax(0,1fr)}}table{{display:block;max-width:100%;overflow-x:auto}}}}@media(max-width:480px){{article{{width:100%;margin:0;border-radius:0}}h1{{font-size:31px}}}}</style></head>
<body><article><header><div class="brand"><span class="mark">Z</span>ZubePredict AI</div><div class="eyebrow">A record for repeating and auditing the run</div><h1>Reproducibility Manifest</h1><p>The experiment recipe, software environment and integrity references in one readable document.</p><span class="stamp">VERIFIED &middot; VERSION {payload['report_version']}</span></header><main>
<div class="intro"><strong>Why this document exists</strong>This manifest helps another reviewer understand what would be needed to repeat or audit the experiment. It does not claim that repeating the run on different data will produce the same performance.</div>
<div class="steps"><div class="step"><b>1</b><strong>Identify the run</strong><p>Match the experiment, dataset fingerprint and Constitution version.</p></div><div class="step"><b>2</b><strong>Recreate the setup</strong><p>Use the recorded random seed, model settings and software versions.</p></div><div class="step"><b>3</b><strong>Verify integrity</strong><p>Compare the evidence hash before trusting any reproduced result.</p></div></div>
<section><h2>Run at a glance</h2><p>The key recipe recorded for this experiment.</p><div class="facts"><div class="fact"><span>Selected model</span><strong>{_safe(_label(payload['selected_model']))}</strong></div><div class="fact"><span>Prediction task</span><strong>{_safe(_label(payload['task']))}</strong></div><div class="fact"><span>Target</span><strong>{_safe(_label(payload['target']))}</strong></div><div class="fact"><span>Random seed</span><strong>{_safe(reproducibility.get('random_seed'))}</strong></div><div class="fact"><span>Validation</span><strong>{_safe(_plain_validation(payload['validation_strategy']))}</strong></div><div class="fact"><span>Generated</span><strong>{_safe(_generated_label(payload['generated_at']))}</strong></div></div><div class="policy"><strong>Artifact policy</strong>{_safe(artifact_policy)}</div><div class="warning"><strong>Use limitation</strong>{_safe(payload['intended_use_warning'])}</div></section>
<section><h2>Data and experiment boundaries</h2><p>These values identify the approved data and Constitution without exposing source rows.</p><h3>Excluded fields</h3><div class="chips">{exclusions or '<span class="chip">None recorded</span>'}</div><details><summary>Show experiment identifiers</summary><table><tr><th>Experiment ID</th><td><code>{_safe(payload['experiment_id'])}</code></td></tr><tr><th>Dataset fingerprint</th><td><code>{_safe(payload['dataset_fingerprint'])}</code></td></tr><tr><th>Constitution version</th><td>{_safe(payload['constitution_version'])}</td></tr><tr><th>Evidence hash</th><td><code>{_safe(payload['integrity_reference'])}</code></td></tr></table></details></section>
<section><h2>Technical recipe</h2><p>Open these sections only when reproducing or auditing the run.</p><details><summary>Show software versions</summary><table>{software_rows or '<tr><td>Not recorded</td></tr>'}</table></details><details><summary>Show selected model settings</summary><table>{parameter_rows or '<tr><td>Not recorded</td></tr>'}</table></details></section></main><footer>This manifest is derived from the immutable Evidence Envelope. Delivery channels must not rewrite its values.</footer></article></body></html>"""
    return document.encode("utf-8")


def _pdf_report(payload: dict[str, Any]) -> bytes:
    output = io.BytesIO()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ZPTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=28,
        leading=32,
        textColor=colors.white,
        alignment=TA_LEFT,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ZPSubtitle",
        parent=styles["BodyText"],
        fontSize=11,
        leading=16,
        textColor=colors.HexColor("#DCEAFF"),
    )
    heading_style = ParagraphStyle(
        "ZPHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        textColor=_NAVY,
        spaceBefore=10,
        spaceAfter=8,
        keepWithNext=True,
    )
    body_style = ParagraphStyle(
        "ZPBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.4,
        leading=14,
        textColor=_INK,
        spaceAfter=5,
    )
    small_style = ParagraphStyle(
        "ZPSmall",
        parent=body_style,
        fontSize=7.6,
        leading=10,
        textColor=_MUTED,
    )
    label_style = ParagraphStyle(
        "ZPLabel",
        parent=small_style,
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=9,
        textColor=_MUTED,
        uppercase=True,
    )
    metric_style = ParagraphStyle(
        "ZPMetric",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=21,
        textColor=_NAVY,
        spaceAfter=0,
    )
    hash_style = ParagraphStyle(
        "ZPHash",
        parent=small_style,
        fontName="Courier",
        fontSize=6.7,
        leading=9,
        wordWrap="CJK",
    )

    def paragraph(value: Any, style: ParagraphStyle = body_style) -> Paragraph:
        return Paragraph(_safe(value), style)

    def section(title: str) -> list[Any]:
        return [Spacer(1, 4 * mm), Paragraph(_safe(title), heading_style)]

    def key_value_table(rows: list[tuple[Any, Any]]) -> Table:
        table = Table(
            [[paragraph(key, label_style), paragraph(value)] for key, value in rows],
            colWidths=[47 * mm, 116 * mm],
            hAlign="LEFT",
        )
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (0, 0), (0, -1), _SURFACE),
                    ("BOX", (0, 0), (-1, -1), 0.6, _BORDER),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, _BORDER),
                    ("LEFTPADDING", (0, 0), (-1, -1), 9),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        return table

    primary_name = str(payload["primary_metric"])
    primary = _primary_stats(payload)
    hero = Table(
        [[Paragraph("ZUBEPREDICT AI", ParagraphStyle("Brand", parent=label_style, textColor=colors.white)), ""],
         [Paragraph("Evidence Report", title_style), ""],
         [Paragraph("Verified model evaluation · EyeCare Evidence Card", subtitle_style), ""]],
        colWidths=[130 * mm, 33 * mm],
    )
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), _NAVY),
                ("SPAN", (0, 0), (1, 0)),
                ("SPAN", (0, 1), (1, 1)),
                ("SPAN", (0, 2), (1, 2)),
                ("LEFTPADDING", (0, 0), (-1, -1), 16),
                ("RIGHTPADDING", (0, 0), (-1, -1), 16),
                ("TOPPADDING", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 2), (-1, 2), 16),
            ]
        )
    )
    story: list[Any] = [hero, Spacer(1, 7 * mm)]
    story.extend(
        [
            Paragraph("What this result says", heading_style),
            Table(
                [[paragraph(_executive_summary(payload))]],
                colWidths=[163 * mm],
                style=TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), _PALE_BLUE),
                        ("LINEBEFORE", (0, 0), (0, -1), 4, _BLUE),
                        ("BOX", (0, 0), (-1, -1), 0.4, _BORDER),
                        ("LEFTPADDING", (0, 0), (-1, -1), 13),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 13),
                        ("TOPPADDING", (0, 0), (-1, -1), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 11),
                    ]
                ),
            ),
            Spacer(1, 5 * mm),
        ]
    )
    cards = Table(
        [
            [paragraph("SELECTED MODEL", label_style), paragraph(_label(primary_name), label_style), paragraph("TASK", label_style)],
            [paragraph(_label(payload["selected_model"]), metric_style), paragraph(primary["value"], metric_style), paragraph(_label(payload["task"]), metric_style)],
        ],
        colWidths=[58 * mm, 47 * mm, 58 * mm],
    )
    cards.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), _SURFACE),
                ("BOX", (0, 0), (-1, -1), 0.5, _BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, _BORDER),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(cards)
    story.extend(section("Important use limitation"))
    warning = Table([[paragraph(payload["intended_use_warning"])]], colWidths=[163 * mm])
    warning.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), _PALE_AMBER),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#E5C16C")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.append(warning)
    story.extend(section("Study design"))
    fields = [
        ("Constitution version", payload["constitution_version"]),
        ("Prediction target", _label(payload["target"])),
        ("Exclusions", ", ".join(payload["exclusions"]) or "None recorded"),
        ("Validation strategy", payload["validation_strategy"]),
        ("Primary measure", f"{_label(primary_name)} — {_METRIC_GUIDANCE.get(primary_name, 'Main validation measure.')}"),
        ("95% confidence interval", primary["confidence"]),
        ("Standard deviation", primary["spread"]),
    ]
    story.append(key_value_table(fields))

    story.extend([PageBreak(), *section("Model leaderboard")])
    leaderboard_data: list[list[Any]] = [
        [paragraph("RANK", label_style), paragraph("MODEL", label_style), paragraph(_label(primary_name), label_style), paragraph("STATUS", label_style)]
    ]
    for rank, item in enumerate(payload["leaderboard"], start=1):
        if not isinstance(item, dict):
            continue
        metrics = item.get("metrics", {})
        value = metrics.get(primary_name) if isinstance(metrics, dict) else None
        leaderboard_data.append(
            [paragraph(rank), paragraph(_label(item.get("model_name"))), paragraph(_metric_stats(value)["value"]), paragraph(_label(item.get("status", "completed")))]
        )
    leaderboard = Table(leaderboard_data, colWidths=[14 * mm, 83 * mm, 36 * mm, 30 * mm], repeatRows=1)
    leaderboard.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _SURFACE]),
                ("BOX", (0, 0), (-1, -1), 0.5, _BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, _BORDER),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(leaderboard)

    story.extend(section("Supporting measures"))
    metric_data: list[list[Any]] = [
        [paragraph("MEASURE", label_style), paragraph("MEAN", label_style), paragraph("STD. DEV.", label_style), paragraph("95% INTERVAL", label_style)]
    ]
    for name, value in _winner_metrics(payload).items():
        if name == primary_name:
            continue
        stats = _metric_stats(value)
        metric_data.append([paragraph(_label(name)), paragraph(stats["value"]), paragraph(stats["spread"]), paragraph(stats["confidence"])])
    if len(metric_data) == 1:
        metric_data.append([paragraph("No secondary measures recorded"), "", "", ""])
    metric_table = Table(metric_data, colWidths=[58 * mm, 28 * mm, 31 * mm, 46 * mm], repeatRows=1)
    metric_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _SURFACE]),
                ("BOX", (0, 0), (-1, -1), 0.5, _BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, _BORDER),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(metric_table)

    analysis = payload.get("calibration_error_analysis", {})
    calibration = analysis.get("calibration", {}) if isinstance(analysis, dict) else {}
    threshold = analysis.get("threshold_analysis", {}) if isinstance(analysis, dict) else {}
    error_summary = analysis.get("error_analysis_summary", {}) if isinstance(analysis, dict) else {}
    story.extend(section("Calibration and error review"))
    story.append(
        key_value_table(
            [
                ("Brier score", _number(calibration.get("brier_score"))),
                ("Expected calibration error", _number(calibration.get("expected_calibration_error"))),
                ("Recommended threshold", _number(threshold.get("recommended_threshold"))),
                ("Threshold recommendation", threshold.get("recommendation_basis") or "Not recorded"),
                ("Calibration bins evaluated", len(calibration.get("bins", []))),
                ("Error-analysis segments", _number(error_summary.get("segment_count"))),
                ("Recorded plots", ", ".join(_label(item) for item in error_summary.get("plot_ids", [])) or "None recorded"),
                ("Protected fields skipped", ", ".join(str(item) for item in error_summary.get("protected_columns_skipped", [])) or "None recorded"),
            ]
        )
    )

    story.extend([PageBreak(), *section("Limitations and warnings")])
    limitation_items = _list_items(payload["limitations"]) or ["No additional limitations were recorded."]
    for item in limitation_items:
        story.append(Paragraph(f"&bull;&nbsp; {_safe(item)}", body_style))
    for item in _list_items(payload["warnings"]):
        story.append(Paragraph(f"&bull;&nbsp; {_safe(item)}", body_style))

    reproducibility = payload.get("reproducibility", {})
    story.extend(section("Reproducibility"))
    story.append(
        key_value_table(
            [
                ("Random seed", _number(reproducibility.get("random_seed"))),
                ("Generated", _generated_label(payload["generated_at"])),
                ("Report version", payload["report_version"]),
                ("Validation strategy", payload["validation_strategy"]),
            ]
        )
    )
    software_rows = [("Software · " + _label(name), version) for name, version in reproducibility.get("software_versions", {}).items()]
    if software_rows:
        story.extend([Spacer(1, 3 * mm), key_value_table(software_rows)])
    hyperparameter_rows = [("Parameter · " + _label(name), _number(value)) for name, value in reproducibility.get("winner_hyperparameters", {}).items()]
    if hyperparameter_rows:
        story.extend([Spacer(1, 3 * mm), Paragraph("Selected model settings", heading_style), key_value_table(hyperparameter_rows)])

    story.extend(section("Integrity and traceability"))
    integrity = Table(
        [
            [paragraph("EXPERIMENT ID", label_style), paragraph(payload["experiment_id"], hash_style)],
            [paragraph("DATASET FINGERPRINT", label_style), paragraph(payload["dataset_fingerprint"], hash_style)],
            [paragraph("CONSTITUTION VERSION", label_style), paragraph(payload["constitution_version"], hash_style)],
            [paragraph("EVIDENCE HASH", label_style), paragraph(payload["integrity_reference"], hash_style)],
        ],
        colWidths=[47 * mm, 116 * mm],
    )
    integrity.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), _PALE_GREEN),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#9ACFBE")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9DDD1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(KeepTogether([integrity]))

    def draw_page(canvas: Canvas, document: Any) -> None:
        canvas.saveState()
        width, height = A4
        canvas.setStrokeColor(_BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(18 * mm, 13 * mm, width - 18 * mm, 13 * mm)
        canvas.setFillColor(_MUTED)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(18 * mm, 8.5 * mm, "ZubePredict AI · Verified evidence")
        canvas.drawRightString(width - 18 * mm, 8.5 * mm, f"Page {document.page}")
        canvas.restoreState()

    SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=19 * mm,
        title="ZubePredict AI Evidence Report",
        author="ZubePredict AI",
        subject="Verified model evaluation evidence",
        invariant=1,
    ).build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return output.getvalue()


def _prediction_rows(result: Any) -> list[dict[str, Any]]:
    for attribute in ("out_of_fold_predictions", "assignments", "forecast"):
        values = getattr(result, attribute, None)
        if values:
            return [
                item.model_dump(mode="json") if hasattr(item, "model_dump") else dict(item)
                for item in values
            ]
    return []


def _prediction_csv(evidence: EvidenceEnvelope, rows: list[dict[str, Any]]) -> bytes:
    metadata = {
        "experiment_id": str(evidence.experiment_id),
        "dataset_fingerprint": evidence.dataset_fingerprint,
        "constitution_version": evidence.constitution_version,
        "evidence_hash": evidence.evidence_hash,
    }
    fieldnames = [*metadata, *sorted({key for row in rows for key in row})]
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({**metadata, **row})
    return output.getvalue().encode("utf-8-sig")


def _prediction_xlsx(evidence: EvidenceEnvelope, rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    predictions = workbook.active
    predictions.title = "Predictions"
    headers = sorted({key for row in rows for key in row})
    predictions.append(headers)
    for row in rows:
        predictions.append(
            [
                json.dumps(row.get(key), default=str)
                if isinstance(row.get(key), (dict, list))
                else row.get(key)
                for key in headers
            ]
        )
    header_fill = PatternFill("solid", fgColor="142B4A")
    accent_fill = PatternFill("solid", fgColor="EAF3FF")
    for cell in predictions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    predictions.freeze_panes = "A2"
    predictions.auto_filter.ref = predictions.dimensions
    predictions.sheet_view.showGridLines = False
    for index, heading in enumerate(headers, start=1):
        observed = [len(str(heading)), *[len(str(row.get(heading, ""))) for row in rows[:200]]]
        predictions.column_dimensions[get_column_letter(index)].width = min(max(observed) + 3, 36)
    metadata = workbook.create_sheet("Evidence metadata")
    metadata.append(["Evidence field", "Recorded value"])
    for key, value in (
        ("experiment_id", str(evidence.experiment_id)),
        ("dataset_fingerprint", evidence.dataset_fingerprint),
        ("constitution_version", evidence.constitution_version),
        ("evidence_hash", evidence.evidence_hash),
        ("task", evidence.task_type),
        ("target", evidence.target),
        ("validation_strategy", evidence.validation_strategy),
        ("primary_metric", evidence.primary_metric),
        ("selected_model", evidence.winner),
        ("intended_use_warning", evidence.intended_use_warning),
    ):
        metadata.append(
            [key, json.dumps(value, default=str) if isinstance(value, (dict, list)) else value]
        )
    for cell in metadata[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row in metadata.iter_rows(min_row=2):
        row[0].fill = accent_fill
        row[0].font = Font(bold=True, color="142B4A")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    metadata.column_dimensions["A"].width = 30
    metadata.column_dimensions["B"].width = 92
    metadata.freeze_panes = "A2"
    metadata.sheet_view.showGridLines = False

    readme = workbook.create_sheet("Read me", 0)
    readme.append(["ZubePredict AI prediction export"])
    readme.append(
        [
            "These rows are generated predictions from the recorded validation run. "
            "They are decision-support/research outputs unless independently validated."
        ]
    )
    readme.append(["Selected model", evidence.winner or "Not recorded"])
    readme.append(["Primary measure", _label(evidence.primary_metric)])
    readme.append(["Evidence hash", evidence.evidence_hash])
    readme["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    readme["A1"].fill = header_fill
    readme.merge_cells("A1:B1")
    readme.merge_cells("A2:B2")
    readme["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.row_dimensions[2].height = 45
    readme.column_dimensions["A"].width = 30
    readme.column_dimensions["B"].width = 92
    readme.sheet_view.showGridLines = False
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_report_bundle(evidence: EvidenceEnvelope, result: Any) -> list[GeneratedReport]:
    payload = _common_payload(evidence)
    artifact_policy = (
        "Generated once from the verified evidence envelope; "
        "channel delivery must not rewrite metrics."
    )
    reports = [
        _report(
            "evidence",
            "zubepredict-evidence-envelope.json",
            "application/json",
            _json_bytes(evidence.model_dump(mode="json")),
        ),
        _report(
            "evidence_card",
            "zubepredict-eyecare-evidence-card.html",
            "text/html; charset=utf-8",
            _html_evidence_card(payload),
        ),
        _report(
            "html",
            "zubepredict-evidence-report.html",
            "text/html; charset=utf-8",
            _html_report(payload),
        ),
        _report("pdf", "zubepredict-evidence-report.pdf", "application/pdf", _pdf_report(payload)),
        _report(
            "model_card",
            "zubepredict-model-card.html",
            "text/html; charset=utf-8",
            _model_card(payload),
        ),
        _report(
            "reproducibility_manifest",
            "zubepredict-reproducibility-manifest.html",
            "text/html; charset=utf-8",
            _html_reproducibility_manifest(payload, artifact_policy),
        ),
    ]
    rows = _prediction_rows(result)
    if rows:
        reports.extend(
            [
                _report(
                    "predictions_csv",
                    "zubepredict-predictions.csv",
                    "text/csv; charset=utf-8",
                    _prediction_csv(evidence, rows),
                ),
                _report(
                    "predictions_xlsx",
                    "zubepredict-predictions.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    _prediction_xlsx(evidence, rows),
                ),
            ]
        )
    return reports


def required_report_types(*, predictions_available: bool) -> set[str]:
    base = {"evidence", "evidence_card", "html", "pdf", "model_card", "reproducibility_manifest"}
    return base | ({"predictions_csv", "predictions_xlsx"} if predictions_available else set())
