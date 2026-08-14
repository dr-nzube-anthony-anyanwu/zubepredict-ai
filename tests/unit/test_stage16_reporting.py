from __future__ import annotations

import hashlib
import io
import json
from datetime import UTC, datetime
from types import SimpleNamespace
from typing import Any
from uuid import UUID, uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from zubepredict_api.routes import hermes as hermes_routes
from zubepredict_api.security.hermes import TrustedHermesPrincipal
from zubepredict_core.evidence import (
    EvidenceEnvelope,
    build_evidence_envelope,
    grounded_or_fallback_narrative,
    verify_evidence_envelope,
)
from zubepredict_core.reporting import generate_report_bundle, required_report_types
from zubepredict_core.repositories.models import ExperimentRecord, ReportRecord
from zubepredict_core.shared.schemas import OutOfFoldPrediction

from integrations.hermes.plugin.zubepredict import report_delivery

OWNER = UUID("11111111-1111-4111-8111-111111111111")
EXPERIMENT = UUID("22222222-2222-4222-8222-222222222222")
PROJECT = UUID("33333333-3333-4333-8333-333333333333")
DATASET = UUID("44444444-4444-4444-8444-444444444444")
GENERATED_AT = datetime(2026, 8, 14, 12, 0, tzinfo=UTC)


def _evidence() -> EvidenceEnvelope:
    return build_evidence_envelope(
        experiment_id=EXPERIMENT,
        dataset_fingerprint="a" * 64,
        constitution_version=3,
        result_summary={
            "task_type": "binary_classification",
            "target": "readmitted",
            "validation_strategy": "stratified_cross_validation",
            "primary_metric": "pr_auc",
            "winner": "logistic_regression",
            "leaderboard": [
                {
                    "model_name": "logistic_regression",
                    "metrics": {
                        "pr_auc": {"mean": 0.81},
                        "roc_auc": {"mean": 0.84},
                    },
                    "hyperparameters": {"C": 1.0},
                }
            ],
            "calibration": {"brier_score": 0.12},
            "error_analysis_summary": {"plot_ids": ["confusion_matrix"]},
            "random_seed": 42,
            "software_versions": {"python": "3.11.0"},
        },
        warnings=["Synthetic test data only."],
        constitution={
            "exclusions": ["patient_id"],
            "intended_use_warning": "Decision support and research only.",
        },
        generated_at=GENERATED_AT,
    )


def _result() -> SimpleNamespace:
    return SimpleNamespace(
        out_of_fold_predictions=[
            OutOfFoldPrediction(
                row_index="0",
                fold=1,
                actual=0,
                predicted=0,
                positive_probability=0.15,
            )
        ]
    )


def test_bundle_contains_required_authoritative_artifacts_and_fields() -> None:
    evidence = _evidence()
    reports = generate_report_bundle(evidence, _result())
    by_type = {item.report_type: item for item in reports}

    assert set(by_type) == required_report_types(predictions_available=True)
    assert {item.report_version for item in reports} == {3}
    assert by_type["pdf"].content.startswith(b"%PDF")
    assert b"ZubePredict AI Evidence Report" in by_type["html"].content
    assert b"What this result says" in by_type["html"].content
    assert b"Model leaderboard" in by_type["html"].content
    assert b"What was predicted?" in by_type["html"].content
    assert b"What the score does not mean" in by_type["html"].content
    assert b"Show integrity and traceability details" in by_type["html"].content
    assert b"<pre>" not in by_type["html"].content
    assert str(EXPERIMENT).encode() in by_type["html"].content
    assert evidence.dataset_fingerprint.encode() in by_type["model_card"].content
    assert by_type["model_card"].filename == "zubepredict-model-card.html"
    assert by_type["model_card"].content_type == "text/html; charset=utf-8"
    assert b"What is a Model Card?" in by_type["model_card"].content
    assert b"Responsible use" in by_type["model_card"].content
    assert b"Show software versions" in by_type["model_card"].content
    assert all(item.sha256 == hashlib.sha256(item.content).hexdigest() for item in reports)

    card = by_type["evidence_card"]
    assert card.filename == "zubepredict-eyecare-evidence-card.html"
    assert card.content_type == "text/html; charset=utf-8"
    assert b"EyeCare Evidence Card" in card.content
    assert b"Start here." in card.content
    assert b"How to read this card" in card.content
    assert b"Show technical traceability details" in card.content
    assert b"Logistic Regression" in card.content
    assert evidence.evidence_hash.encode() in card.content
    for human_facing in (by_type["html"], by_type["model_card"], card):
        assert b"\xc3\x82" not in human_facing.content
        assert b"\xc3\xa2\xe2\x82\xac" not in human_facing.content

    manifest = json.loads(by_type["reproducibility_manifest"].content)
    assert manifest["constitution_version"] == 3
    assert manifest["task"] == "binary_classification"
    assert manifest["target"] == "readmitted"
    assert manifest["exclusions"] == ["patient_id"]
    assert manifest["primary_metric"] == "pr_auc"
    assert "roc_auc" in manifest["secondary_metrics"]
    assert manifest["selected_model"] == "logistic_regression"
    assert manifest["calibration_error_analysis"]["calibration"]["brier_score"] == 0.12
    assert manifest["reproducibility"]["random_seed"] == 42
    assert by_type["evidence"].content.startswith(b"{\n")


def test_prediction_exports_contain_integrity_metadata_without_source_features() -> None:
    evidence = _evidence()
    by_type = {item.report_type: item for item in generate_report_bundle(evidence, _result())}

    csv_text = by_type["predictions_csv"].content.decode("utf-8-sig")
    assert "dataset_fingerprint" in csv_text
    assert "evidence_hash" in csv_text
    assert "patient_id" not in csv_text

    workbook = load_workbook(io.BytesIO(by_type["predictions_xlsx"].content))
    assert workbook.sheetnames == ["Read me", "Predictions", "Evidence metadata"]
    assert workbook["Predictions"].freeze_panes == "A2"
    assert workbook["Predictions"]["A1"].fill.fgColor.rgb == "00142B4A"
    metadata = dict(workbook["Evidence metadata"].iter_rows(values_only=True))
    assert metadata["experiment_id"] == str(EXPERIMENT)
    assert metadata["evidence_hash"] == evidence.evidence_hash


def test_evidence_tampering_and_conflicting_narrative_are_rejected() -> None:
    evidence = _evidence()
    assert verify_evidence_envelope(evidence)
    tampered = evidence.model_copy(update={"winner": "random_forest"})
    assert not verify_evidence_envelope(tampered)
    assert "random_forest" not in grounded_or_fallback_narrative(
        evidence, "random_forest scored 99.99"
    )


class FakeBucket:
    def __init__(self, content: bytes) -> None:
        self.content = content
        self.signed_paths: list[tuple[str, int]] = []

    def download(self, _path: str) -> bytes:
        return self.content

    def create_signed_url(self, path: str, ttl: int) -> dict[str, str]:
        self.signed_paths.append((path, ttl))
        return {"signedURL": f"https://storage.example.test/object?token={uuid4().hex}"}


class FakeStorage:
    def __init__(self, bucket: FakeBucket) -> None:
        self.bucket = bucket

    def from_(self, _name: str) -> FakeBucket:
        return self.bucket


class FakeReports:
    def __init__(self, reports: list[ReportRecord]) -> None:
        self.reports = reports

    def list_for_experiment(self, _experiment_id: UUID) -> list[ReportRecord]:
        return self.reports


class FakeAudit:
    def __init__(self) -> None:
        self.records: list[dict[str, Any]] = []

    def record(self, **values: Any) -> None:
        self.records.append(values)


def _experiment(status: str = "completed") -> ExperimentRecord:
    return ExperimentRecord(
        id=EXPERIMENT,
        owner_id=OWNER,
        project_id=PROJECT,
        dataset_id=DATASET,
        status=status,
        result_summary={"evidence_hash": _evidence().evidence_hash},
    )


def _delivery_setup(
    monkeypatch: pytest.MonkeyPatch,
    *,
    content: bytes,
    experiment: ExperimentRecord | None = None,
    sha256: str | None = None,
    evidence_hash: str | None = None,
) -> tuple[ReportRecord, FakeBucket, FakeAudit]:
    report = ReportRecord(
        id=uuid4(),
        owner_id=OWNER,
        experiment_id=EXPERIMENT,
        report_type="evidence",
        report_version=1,
        storage_path=f"{OWNER}/{EXPERIMENT}/reports/v1/zubepredict-evidence-envelope.json",
        filename="zubepredict-evidence-envelope.json",
        content_type="application/json",
        size_bytes=len(content),
        sha256=sha256 or hashlib.sha256(content).hexdigest(),
        evidence_hash=evidence_hash or _evidence().evidence_hash,
    )
    bucket = FakeBucket(content)
    audit = FakeAudit()
    repositories = SimpleNamespace(
        experiments=SimpleNamespace(get=lambda _id: experiment or _experiment()),
        reports=FakeReports([report]),
        audit_logs=audit,
    )
    monkeypatch.setattr(hermes_routes, "_repositories", lambda _principal: repositories)
    monkeypatch.setattr(
        hermes_routes,
        "create_service_session",
        lambda *_args: SimpleNamespace(client=SimpleNamespace(storage=FakeStorage(bucket))),
    )
    monkeypatch.setattr(
        hermes_routes,
        "get_settings",
        lambda: SimpleNamespace(
            supabase_artifacts_bucket="artifacts",
            hermes_telegram_report_ttl_seconds=300,
        ),
    )
    return report, bucket, audit


def test_web_telegram_and_api_receive_the_same_artifact_identity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    evidence = _evidence()
    content = evidence.model_dump_json().encode()
    report, bucket, audit = _delivery_setup(monkeypatch, content=content)

    responses = [
        hermes_routes.get_report_reference(
            EXPERIMENT,
            "evidence",
            TrustedHermesPrincipal(OWNER, channel, channel=channel),
        )
        for channel in ("web", "telegram", "api")
    ]

    assert {item["report_id"] for item in responses} == {str(report.id)}
    assert {item["sha256"] for item in responses} == {report.sha256}
    assert {item["evidence_hash"] for item in responses} == {evidence.evidence_hash}
    assert bucket.signed_paths == [(report.storage_path, 300)] * 3
    assert len(audit.records) == 3


def test_unfinished_cross_owner_and_tampered_reports_fail_safely(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    evidence = _evidence()
    content = evidence.model_dump_json().encode()
    _delivery_setup(monkeypatch, content=content, experiment=_experiment("running"))
    principal = TrustedHermesPrincipal(OWNER, "test", channel="api")
    with pytest.raises(HTTPException) as unfinished:
        hermes_routes.get_report_reference(EXPERIMENT, "evidence", principal)
    assert unfinished.value.status_code == 409

    repositories = SimpleNamespace(
        experiments=SimpleNamespace(get=lambda _id: None),
        reports=FakeReports([]),
        audit_logs=FakeAudit(),
    )
    monkeypatch.setattr(hermes_routes, "_repositories", lambda _principal: repositories)
    with pytest.raises(HTTPException) as cross_owner:
        hermes_routes.get_report_reference(EXPERIMENT, "evidence", principal)
    assert cross_owner.value.status_code == 404

    _delivery_setup(monkeypatch, content=content + b"tampered", sha256="f" * 64)
    with pytest.raises(HTTPException) as tampered:
        hermes_routes.get_report_reference(EXPERIMENT, "evidence", principal)
    assert tampered.value.status_code == 409
    assert "integrity" in str(tampered.value.detail).lower()

    _delivery_setup(monkeypatch, content=content, evidence_hash="b" * 64)
    with pytest.raises(HTTPException) as detached:
        hermes_routes.get_report_reference(EXPERIMENT, "evidence", principal)
    assert detached.value.status_code == 409


def test_expired_telegram_delivery_is_not_rendered(monkeypatch: pytest.MonkeyPatch) -> None:
    report_delivery._PENDING_BY_SESSION.clear()
    clock = iter([100.0, 100.0, 401.0])
    monkeypatch.setattr(report_delivery.time, "monotonic", lambda: next(clock))
    report_delivery.remember_report_delivery(
        "owner-session",
        json.dumps(
            {
                "ok": True,
                "data": {
                    "download_url": "https://storage.example.test/report?token=signed",
                    "download_filename": "zubepredict-evidence-report.pdf",
                    "expires_in_seconds": 300,
                },
            }
        ),
    )

    assert report_delivery.render_pending_report("ignored", "owner-session", "telegram") is None
